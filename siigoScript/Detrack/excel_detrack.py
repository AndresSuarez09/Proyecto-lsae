# excel_detrack.py

import os
import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

CARPETA_SALIDA = "resultados"
NOMBRE_BASE = "ordenes_detrack.xlsx"

def _obtener_nombre_incremental():
    """
    Genera un nombre incremental para el archivo Excel.
    Ejemplo: ordenes_detrack.xlsx, ordenes_detrack(1).xlsx, ordenes_detrack(2).xlsx...
    """
    patron = os.path.join(CARPETA_SALIDA, "ordenes_detrack*.xlsx")
    archivos = glob.glob(patron)
    if not archivos:
        return os.path.join(CARPETA_SALIDA, NOMBRE_BASE)
    else:
        # Buscar el mayor índice usado
        indices = []
        for archivo in archivos:
            nombre = os.path.basename(archivo)
            if "(" in nombre and ")" in nombre:
                try:
                    idx = int(nombre.split("(")[1].split(")")[0])
                    indices.append(idx)
                except:
                    pass
        nuevo_idx = max(indices) + 1 if indices else 1
        return os.path.join(CARPETA_SALIDA, f"ordenes_detrack({nuevo_idx}).xlsx")

def generar_excel_detrack(ordenes):
    """
    Genera un Excel con las órdenes consultadas en Detrack.
    """
    if not ordenes:
        print("🚫 No se encontraron órdenes en Detrack.")
        return None

    os.makedirs(CARPETA_SALIDA, exist_ok=True)

    # ✅ Seleccionar y normalizar campos relevantes
    normalizadas = []
    for job in ordenes:
        limpio = {
            "DO Number": job.get("do_number", ""),
            "Fecha": job.get("date", ""),
            "Cliente": job.get("deliver_to_collect_from", ""),
            "Dirección": job.get("address", ""),
            "Estado": job.get("status", ""),
            "Detrack #": job.get("id", ""),
            "Tracking Link": job.get("tracking_link", ""),
            "Observaciones": job.get("remarks", ""),
        }
        normalizadas.append(limpio)

    # Crear DataFrame con datos normalizados
    df = pd.DataFrame(normalizadas)

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Órdenes Detrack"

    # Encabezados en rojo con letras blancas
    encabezado_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    encabezado_font = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center")

    # Escribir DataFrame en hoja
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(row)
        if r_idx == 1:  # primera fila = encabezados
            for cell in ws[r_idx]:
                cell.fill = encabezado_fill
                cell.font = encabezado_font
                cell.alignment = align_center

    # Alternar colores de filas
    fill_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for r_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        if r_idx % 2 == 0:
            for cell in row:
                cell.fill = fill_gray

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # Guardar archivo con nombre incremental
    ruta_salida = _obtener_nombre_incremental()
    wb.save(ruta_salida)
    print(f"✅ Excel generado: {ruta_salida}")
    return ruta_salida