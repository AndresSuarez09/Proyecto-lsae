# excel_generator.py
import config
import os
import glob
import re
import pandas as pd
import requests
import time
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

clientes_cache = {}
consultas_realizadas = 0

# 🔧 Resolver cliente por identificación (incluye dirección)
def resolver_cliente_por_identificacion(identificacion, token, clientes_resueltos=None):
    global consultas_realizadas

    if not identificacion or not identificacion.strip() or identificacion == "222222222222":
        return f"ID: {identificacion}", None, None

    if clientes_resueltos and identificacion in clientes_resueltos:
        nombre = clientes_resueltos[identificacion]
        clientes_cache[identificacion] = (nombre, identificacion, None)
        return nombre, identificacion, None

    if identificacion in clientes_cache:
        return clientes_cache[identificacion]

    if consultas_realizadas >= 95:
        time.sleep(60)
        consultas_realizadas = 0

    url = f"https://api.siigo.com/v1/customers?identification={identificacion}"
    headers = {
        "Authorization": f"{token}",
        "Content-Type": "application/json",
        "Partner-Id": "Lubrisol"
    }

    try:
        response = requests.get(url, headers=headers)
        consultas_realizadas += 1
        data = response.json()

        if isinstance(data, dict) and "results" in data and isinstance(data["results"], list) and data["results"]:
            cliente = data["results"][0]
            nombre_raw = cliente.get("name")
            nombre = " ".join(nombre_raw) if isinstance(nombre_raw, list) else nombre_raw
            nit = cliente.get("identification", identificacion)

            # Dirección completa
            direccion_obj = cliente.get("address", {})
            direccion_parts = []
            if direccion_obj.get("address"):
                direccion_parts.append(direccion_obj.get("address").strip())
            if direccion_obj.get("city", {}).get("city_name"):
                direccion_parts.append(direccion_obj.get("city", {}).get("city_name").strip())
            if direccion_obj.get("postal_code"):
                direccion_parts.append(direccion_obj.get("postal_code").strip())
            direccion = ", ".join(direccion_parts) if direccion_parts else None

            if nombre:
                clientes_cache[identificacion] = (nombre, nit, direccion)
                return nombre, nit, direccion

        clientes_cache[identificacion] = (f"ID: {identificacion}", None, None)
        return f"ID: {identificacion}", None, None

    except Exception:
        clientes_cache[identificacion] = (f"ID: {identificacion}", None, None)
        return f"ID: {identificacion}", None, None
# 🔧 Obtener diccionario de vendedores
def obtener_diccionario_vendedores(token):
    url = "https://api.siigo.com/v1/users"
    headers = {
        "Authorization": f"{token}",
        "Content-Type": "application/json",
        "Partner-Id": "Lubrisol"
    }

    try:
        response = requests.get(url, headers=headers)
        data = response.json()
        vendedores = {}
        for usuario in data.get("results", data):
            seller_id = usuario.get("id")
            nombre = f"{usuario.get('first_name', '')} {usuario.get('last_name', '')}".strip()
            vendedores[seller_id] = {
                "nombre": nombre,
                "email": usuario.get("email", ""),
                "identification": usuario.get("identification", "")
            }
        return vendedores
    except Exception:
        return {}

# 🔧 Aplicar formato estético a un Excel (incluye relleno alterno)
def aplicar_formato_excel(ruta, columnas_ordenadas):
    wb = load_workbook(ruta)
    ws = wb.active

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    fill_gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Encabezados
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Ajuste de ancho de columnas
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)

    # Formato monetario y relleno alterno
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(columnas_ordenadas)), start=2):
        if idx % 2 == 0:  # filas pares → gris claro
            for cell in row:
                cell.fill = fill_gray
        for cell in row:
            if cell.column_letter in ["G", "H"]:  # columnas Total y Saldo
                cell.number_format = '"$"#,##0.00'

    wb.save(ruta)
    print(f"✅ Formato aplicado: {ruta}")
# 🔧 Generar Excel
def generar_excel(facturas, token, clientes_resueltos=None):
    if not facturas:
        print("⚠️ No hay facturas para generar el Excel.")
        return None

    if clientes_resueltos is None:
        clientes_resueltos = {}

    vendedores_diccionario = obtener_diccionario_vendedores(token)
    os.makedirs(config.CARPETA_SALIDA, exist_ok=True)

    base_name = config.NOMBRE_ARCHIVO_EXCEL
    nombre_sin_ext, ext = os.path.splitext(base_name)
    if not ext:
        ext = ".xlsx"
        base_name = f"{base_name}{ext}"

    ruta_fija = os.path.join(config.CARPETA_SALIDA, base_name)
    print(f"📁 Generando archivo Excel (borrador) en: {ruta_fija}")

    registros = []
    for f in facturas:
        if isinstance(f, dict):
            productos, codigos = [], []
            for item in f.get("items", []):
                if isinstance(item, dict):
                    nombre = item.get("description", "Sin nombre")
                    cantidad = item.get("quantity", 0)
                    precio = item.get("price", 0)
                    total = item.get("total", 0)
                    productos.append(f"{nombre} ({cantidad} x {precio:,}) = {total:,}")
                    codigos.append(str(item.get("code", "")))

            # Cliente
            cliente_obj = f.get("customer", {})
            clave_cliente = cliente_obj.get("id") or cliente_obj.get("identification")
            cliente_nombre = clientes_resueltos.get(clave_cliente, "SIN NOMBRE")
            nit_cliente = cliente_obj.get("identification")

            # Dirección: concatenar partes si existen
            direccion_obj = cliente_obj.get("address", {})
            direccion_parts = []
            if direccion_obj.get("address"):
                direccion_parts.append(direccion_obj.get("address").strip())
            if direccion_obj.get("city", {}).get("city_name"):
                direccion_parts.append(direccion_obj.get("city", {}).get("city_name").strip())
            if direccion_obj.get("postal_code"):
                direccion_parts.append(direccion_obj.get("postal_code").strip())
            direccion_final = " ".join(direccion_parts) if direccion_parts else None

            # Si no hay dirección en la factura, consultar API
            if not direccion_final:
                nombre_resuelto, nit_resuelto, direccion_resuelta = resolver_cliente_por_identificacion(nit_cliente, token, clientes_resueltos)
                if nit_resuelto:
                    cliente_nombre = nombre_resuelto
                    nit_cliente = nit_resuelto
                    direccion_final = direccion_resuelta

            # Vendedor
            vendedor_id = f.get("seller")
            vendedor_info = vendedores_diccionario.get(vendedor_id, {})
            vendedor_nombre = vendedor_info.get("nombre", f"ID: {vendedor_id}")
            email_vendedor = vendedor_info.get("email", "")
            id_vendedor = vendedor_info.get("identification", "")

            # Factura
            numero_fv = str(f.get("number")) if f.get("number") else None
            numero_be = numero_fv.replace("FV", "BE") if numero_fv else None

            registros.append({
                "FV": numero_fv,
                "BE": numero_be,
                "Fecha": f.get("date"),
                "company_name": cliente_nombre,
                "deliver_to_collect_from": cliente_nombre,
                "address": direccion_final,
                "city": direccion_obj.get("city", {}).get("city_name") if direccion_obj else None,
                "postal_code": direccion_obj.get("postal_code") if direccion_obj else None,
                "Identificación": nit_cliente,
                "Vendedor": vendedor_nombre,
                "Total": f.get("total"),
                "Saldo": f.get("balance"),
                "Productos": "; ".join(productos),
                "Email vendedor": email_vendedor,
                "ID vendedor": id_vendedor,
                "Código producto": "; ".join(codigos)
            })

    columnas_ordenadas = [
        "FV", "BE", "Fecha", "company_name", "deliver_to_collect_from",
        "address", "city", "postal_code", "Identificación", "Vendedor",
        "Total", "Saldo", "Productos", "Email vendedor", "ID vendedor", "Código producto"
    ]

    df = pd.DataFrame(registros)
    for col in columnas_ordenadas:
        if col not in df.columns:
            df[col] = None
    df = df[columnas_ordenadas]

    # Guardar borrador fijo y aplicar formato
    df.to_excel(ruta_fija, index=False)
    aplicar_formato_excel(ruta_fija, columnas_ordenadas)

    # Crear copia incremental y aplicar formato
    archivos_coincidentes = glob.glob(os.path.join(config.CARPETA_SALIDA, f"{nombre_sin_ext}(*).xlsx"))
    max_index = 0
    regex_indice = re.compile(rf"{re.escape(nombre_sin_ext)}\((\d+)\)\.xlsx$")
    for a in archivos_coincidentes:
        base = os.path.basename(a)
        m = regex_indice.match(base)
        if m:
            idx = int(m.group(1))
            if idx > max_index:
                max_index = idx

    siguiente = max_index + 1
    nuevo_nombre = f"{nombre_sin_ext}({siguiente}).xlsx"
    ruta_incremental = os.path.join(config.CARPETA_SALIDA, nuevo_nombre)
    df.to_excel(ruta_incremental, index=False)
    aplicar_formato_excel(ruta_incremental, columnas_ordenadas)

    print("✅ Archivos Excel generados con formato aplicado.")
    return ruta_fija